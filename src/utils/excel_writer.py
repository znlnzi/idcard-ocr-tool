# -*- coding: utf-8 -*-
"""
Excel文件写入工具
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import os
from ..config.settings import EXCEL_COLUMNS
from .file_handler import FileHandler


class ExcelWriter:
    
    def __init__(self):
        self.file_handler = FileHandler()
        
    def write_results(self, results, output_file):
        """将识别结果写入Excel文件"""
        try:
            # 确保输出目录存在
            self.file_handler.ensure_directory_exists(output_file)
            
            # 创建工作簿
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "身份证信息提取结果"
            
            # 设置样式
            self._setup_styles(worksheet)
            
            # 写入表头
            self._write_header(worksheet)
            
            # 写入数据
            self._write_data(worksheet, results)
            
            # 调整列宽
            self._adjust_column_widths(worksheet)
            
            # 添加统计信息
            self._add_summary(worksheet, results)
            
            # 保存文件
            workbook.save(output_file)
            
            return True, f"Excel文件已保存: {output_file}"
            
        except Exception as e:
            return False, f"保存Excel文件失败: {str(e)}"
            
    def _setup_styles(self, worksheet):
        """设置样式"""
        # 定义样式
        self.header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        self.data_font = Font(name='微软雅黑', size=10)
        self.data_alignment = Alignment(horizontal='left', vertical='center')
        
        self.success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        self.error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.border = thin_border
        
    def _write_header(self, worksheet):
        """写入表头"""
        headers = EXCEL_COLUMNS
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
            
    def _write_data(self, worksheet, results):
        """写入数据"""
        for row, result in enumerate(results, 2):
            # 文件名
            cell = worksheet.cell(row=row, column=1, value=result.get('filename', ''))
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.border
            
            # 姓名
            cell = worksheet.cell(row=row, column=2, value=result.get('name', ''))
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.border
            
            # 民族
            cell = worksheet.cell(row=row, column=3, value=result.get('ethnicity', ''))
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.border
            
            # 识别状态
            status = result.get('status', '')
            cell = worksheet.cell(row=row, column=4, value=status)
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.border
            
            # 根据状态设置背景色
            if status == '成功':
                cell.fill = self.success_fill
            elif status in ['失败', '错误']:
                cell.fill = self.error_fill
                
            # 备注
            cell = worksheet.cell(row=row, column=5, value=result.get('note', ''))
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.border
            
    def _adjust_column_widths(self, worksheet):
        """调整列宽"""
        column_widths = {
            'A': 30,  # 文件名
            'B': 15,  # 姓名
            'C': 15,  # 民族
            'D': 12,  # 识别状态
            'E': 40   # 备注
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
            
    def _add_summary(self, worksheet, results):
        """添加统计信息"""
        if not results:
            return
            
        # 在数据下方添加统计信息
        last_row = len(results) + 1
        summary_start_row = last_row + 3
        
        # 统计数据
        total_count = len(results)
        success_count = sum(1 for r in results if r.get('status') == '成功')
        failed_count = total_count - success_count
        success_rate = (success_count / total_count * 100) if total_count > 0 else 0
        
        # 写入统计信息
        summary_data = [
            ('处理统计', ''),
            ('总文件数', total_count),
            ('成功识别', success_count),
            ('识别失败', failed_count),
            ('成功率', f'{success_rate:.1f}%'),
            ('处理时间', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        for i, (label, value) in enumerate(summary_data):
            row = summary_start_row + i
            
            # 标签
            cell = worksheet.cell(row=row, column=1, value=label)
            if i == 0:  # 标题行
                cell.font = Font(name='微软雅黑', size=11, bold=True)
            else:
                cell.font = self.data_font
            cell.alignment = self.data_alignment
            
            # 值
            if value != '':
                cell = worksheet.cell(row=row, column=2, value=value)
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                
    def create_template_excel(self, output_file):
        """创建Excel模板文件"""
        try:
            self.file_handler.ensure_directory_exists(output_file)
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "身份证信息提取模板"
            
            # 设置样式
            self._setup_styles(worksheet)
            
            # 写入表头
            self._write_header(worksheet)
            
            # 添加示例数据
            example_data = [
                {
                    'filename': 'example_id_card.jpg',
                    'name': '张三',
                    'ethnicity': '汉族',
                    'status': '成功',
                    'note': '识别成功'
                }
            ]
            
            self._write_data(worksheet, example_data)
            self._adjust_column_widths(worksheet)
            
            workbook.save(output_file)
            
            return True, f"模板文件已创建: {output_file}"
            
        except Exception as e:
            return False, f"创建模板文件失败: {str(e)}"